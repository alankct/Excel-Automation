import sys
from openpyxl import Workbook, load_workbook

# Workbook commands

SUPPORTED_FORMATS = (".xlsx", ".xlsm",".xltx", ".xltm")
WB_COMMANDS = ("/wbcreate", "/wbload")

""" 
The workbook commands are always ran first, in order to create or load an excel workbook. These
functions can be called at any time to create or load new workbooks. Returns a workbook, and its
name.
"""
def workbook(wbcommand):
    
    wb = None
    wb_name = None

    match wbcommand:
        case "":
            pass
        case "/wb":
            print(f"Workbook commands: {WB_COMMANDS}")
        case "/wbcreate": # Creates a new workbook
            wb, wb_name =  wbcreate()
        case "/wbload": # Loads an existing workbook
            wb, wb_name = wbload()
        case _:
            print("Invalid /wb command.")
    
    return wb, wb_name


"""
/wbcreate

Creates a new workbook, making sure it's a valid Excel filename, checking to see if it overrides
a current-directory file (while giving the options to re-type a new file name or overrun the file),
and making sure it can be ran by openpyxl. Returns a workbook and its name. 
"""
def wbcreate():
    
    print("TIP: To avoid overriding exissting Excel files, make sure this is an original filename!")
    while True:

        filename = check_filename()
        if filename == None: # User inputted a return command
            return None, None
            
        try:
            # If a file doesn't exist in the directory, this throws an exception and breaks the loop
            load_workbook(filename)

            override = input("This filename already exists, do you want to override it? /yes or /no ")
            if override == "/end":
                print("Program ended")
                sys.exit(0)
            
            if override == "/back":
                return None, None

            if override == "/yes":
                print("Overriding...")
                break
            print("File was not overridden. Creating a new workbook...")

        except:
            break
    
    # Catches any non-functioning file names, specific to Excel and openpyxl
    try:
        wb = Workbook()
        wb.save(filename)
        return wb, filename

    # If an invalid file name is caught, user is prompted to create a new workbook again
    except Exception as e:
        print(e)
        return wbcreate()


"""
/wbload

Loads an existing Excel file from the current or from a specific directory. If the user-inputted file
does not exist, or is not an Excel file, it asks the user for a new filename. Returns the workbook and
its name.
"""
def wbload():
    
    while True:
        
        # Catches any non-functioning file names
        try:
            filename = check_filename()
            if filename == None: # User inputted a return command
                return None, None
            wb = load_workbook(filename)
            wb.save(filename)
            return wb, filename
        
        # If an invalid file name is caught, user is prompted to load a workbook again
        except Exception as e:
            print(e)
            return wbload()


"""
Checks the validity of the Excel filename, by first making sure the user has not typed /end or /back; 
ensuring the program works fluidly. Returns the filename if this file is able to be opened by the program.
"""
def check_filename():
    while True:

        filename = input(f"Enter filename (supported formats are: {SUPPORTED_FORMATS}): ")
            
        if filename in "/end":
            print("Program ended")
            sys.exit(0)

        elif filename in "/back":
            return None

        elif len(filename) <= 5 or filename[-5:] not in SUPPORTED_FORMATS:
            print("This filename is invalid, try again:")
            continue

        else:
            return filename